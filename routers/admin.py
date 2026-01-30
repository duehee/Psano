from __future__ import annotations

from io import BytesIO
from typing import Optional, Dict

from fastapi import APIRouter, Depends, Query, HTTPException, File, UploadFile
from sqlalchemy.orm import Session
from sqlalchemy import text

from database import get_db
from util.utils import iso, now_kst_naive, get_config
from util.constants import MAX_QUESTIONS
from routers.persona import _generate_persona
from routers._store import LOCK, GLOBAL_STATE
from schemas.admin import (
    AdminSessionsResponse, AdminProgressResponse,
    AdminResetRequest, AdminResetResponse,
    AdminPhaseSetRequest, AdminPhaseSetResponse,
    AdminSetCurrentQuestionRequest, AdminSetCurrentQuestionResponse,
    ImportErrorItem, AdminQuestionsImportResponse, AdminSettingsImportResponse,
    AdminPersonalitySetRequest, AdminPersonalitySetResponse, AdminPersonalityGetResponse,
)
from schemas.persona import PersonaGenerateResponse, PersonaGenerateRequest

try:
    from openpyxl import load_workbook
except ImportError:
    raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")


router = APIRouter()


# =========================
# 공통 유틸리티
# =========================

def ensure_psano_state_row(db: Session):
    row = db.execute(text("SELECT id FROM psano_state WHERE id=1")).mappings().first()
    if row:
        return
    db.execute(
        text("""
            INSERT INTO psano_state (id, phase, current_question)
            VALUES (1, 'teach', 1)
        """)
    )


# =========================
# 엑셀 파싱 유틸리티
# =========================

def _cell_str(ws, col: str, row: int) -> str:
    try:
        v = ws[f"{col}{row}"].value
        return "" if v is None else str(v).strip()
    except Exception:
        return ""


def _cell_int(ws, col: str, row: int) -> Optional[int]:
    try:
        v = ws[f"{col}{row}"].value
        if v is None or str(v).strip() == "":
            return None
        return int(v)
    except Exception:
        return None


def _cell_float(ws, col: str, row: int) -> Optional[float]:
    try:
        v = ws[f"{col}{row}"].value
        if v is None or str(v).strip() == "":
            return None
        return float(v)
    except Exception:
        return None


def _get_sheet(wb, preferred_names: list[str]):
    """워크북에서 선호하는 이름의 시트 찾기"""
    name_map = {ws.title.strip().lower(): ws for ws in wb.worksheets}
    for n in preferred_names:
        ws = name_map.get(n.strip().lower())
        if ws is not None:
            return ws
    return None


def _parse_enabled(raw: str) -> Optional[int]:
    s = (raw or "").strip().upper()
    if s in ("Y", "1", "TRUE", "T"):
        return 1
    if s in ("N", "0", "FALSE", "F"):
        return 0
    return None


# =========================
# Questions Import
# =========================

# 컬럼 매핑 (questions 시트)
Q_COL_ID = "A"
Q_COL_VALUE_A = "D"
Q_COL_VALUE_B = "E"
Q_COL_AXIS_KO = "F"
Q_COL_QUESTION = "H"
Q_COL_CHOICE_A = "I"
Q_COL_CHOICE_B = "J"
Q_COL_ENABLED = "K"

AXIS_MAP: Dict[str, str] = {
    "나의 길": "My way",
    "새로움": "Newness",
    "지금 이 순간": "This moment",
    "성장": "growth",
    "함께": "together",
}

def _map_axis_key(axis_ko: str) -> Optional[str]:
    k = (axis_ko or "").strip()
    if not k:
        return None
    if k in AXIS_MAP:
        return AXIS_MAP[k]
    if k in AXIS_MAP.values():
        return k
    return None


def ensure_psano_personality_row(db: Session):
    row = db.execute(text("SELECT id FROM psano_personality WHERE id=1")).mappings().first()
    if row:
        return
    db.execute(text("""
        INSERT INTO psano_personality
            (id, self_direction, conformity, stimulation, security, hedonism, tradition,
             achievement, benevolence, power, universalism)
        VALUES
            (1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
    """))


@router.post("/questions/import", response_model=AdminQuestionsImportResponse)
async def admin_questions_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    POST /admin/questions/import
    xlsx 업로드로 questions upsert.
    """

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = _get_sheet(wb, ["question", "questions"]) or wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid xlsx file: {e}")

    result = AdminQuestionsImportResponse()

    upsert_sql = text("""
        INSERT INTO questions
            (id, axis_key, question_text, choice_a, choice_b, enabled, value_a_key, value_b_key)
        VALUES
            (:id, :axis_key, :question_text, :choice_a, :choice_b, :enabled, :value_a_key, :value_b_key)
        ON DUPLICATE KEY UPDATE
            axis_key = VALUES(axis_key),
            question_text = VALUES(question_text),
            choice_a = VALUES(choice_a),
            choice_b = VALUES(choice_b),
            enabled = VALUES(enabled),
            value_a_key = VALUES(value_a_key),
            value_b_key = VALUES(value_b_key)
    """)

    try:
        for r in range(2, ws.max_row + 1):
            id_val = _cell_int(ws, Q_COL_ID, r)
            q_text = _cell_str(ws, Q_COL_QUESTION, r)

            if id_val is None and q_text == "":
                continue

            result.processed += 1

            axis_ko = _cell_str(ws, Q_COL_AXIS_KO, r)
            axis_key = _map_axis_key(axis_ko)
            choice_a = _cell_str(ws, Q_COL_CHOICE_A, r)
            choice_b = _cell_str(ws, Q_COL_CHOICE_B, r)
            value_a_key = _cell_str(ws, Q_COL_VALUE_A, r)
            value_b_key = _cell_str(ws, Q_COL_VALUE_B, r)
            enabled_raw = _cell_str(ws, Q_COL_ENABLED, r)
            enabled = _parse_enabled(enabled_raw)

            # 필수값 검증
            if id_val is None:
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Invalid or missing ID (A column)"))
                continue
            if not axis_key:
                result.failed += 1
                result.errors.append(ImportErrorItem(
                    row=r, message=f"Invalid axis (F column). Got '{axis_ko}'. Expected: {list(AXIS_MAP.keys())}"
                ))
                continue
            if q_text == "":
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Missing question_text (H column)"))
                continue
            if choice_a == "" or choice_b == "":
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Missing choice_a(I) or choice_b(J)"))
                continue
            if enabled is None:
                result.failed += 1
                result.errors.append(ImportErrorItem(
                    row=r, message=f"Invalid enabled (K column). Got '{enabled_raw}', expected Y/N"
                ))
                continue

            params = {
                "id": id_val,
                "axis_key": axis_key,
                "question_text": q_text,
                "choice_a": choice_a,
                "choice_b": choice_b,
                "enabled": enabled,
                "value_a_key": value_a_key if value_a_key else None,
                "value_b_key": value_b_key if value_b_key else None,
            }

            res = db.execute(upsert_sql, params)
            if res.rowcount == 1:
                result.inserted += 1
            elif res.rowcount == 2:
                result.updated += 1
            else:
                result.unchanged += 1

        db.commit()
        return result

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {e}")


# =========================
# Settings (Growth Stages) Import
# =========================

# 컬럼 매핑 (settings 시트)
S_COL_STAGE_ID = "A"
S_COL_NAME_KR = "B"
S_COL_NAME_EN = "C"
S_COL_MIN_ANSWERS = "D"
S_COL_MAX_ANSWERS = "E"
S_COL_METAPHOR = "F"
S_COL_CERTAINTY = "G"
S_COL_SENT_LEN = "H"
S_COL_EMPATHY = "I"
S_COL_NOTES = "J"


@router.post("/settings/import", response_model=AdminSettingsImportResponse)
async def admin_settings_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    POST /admin/settings/import
    xlsx 업로드로 psano_growth_stages upsert.
    """

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = _get_sheet(wb, ["setting", "settings", "stage", "stages"]) or wb.worksheets[0]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid xlsx file: {e}")

    result = AdminSettingsImportResponse()

    upsert_sql = text("""
        INSERT INTO psano_growth_stages
            (stage_id, stage_name_kr, stage_name_en,
             min_answers, max_answers,
             metaphor_density, certainty,
             sentence_length, empathy_level, notes)
        VALUES
            (:stage_id, :stage_name_kr, :stage_name_en,
             :min_answers, :max_answers,
             :metaphor_density, :certainty,
             :sentence_length, :empathy_level, :notes)
        ON DUPLICATE KEY UPDATE
            stage_name_kr    = VALUES(stage_name_kr),
            stage_name_en    = VALUES(stage_name_en),
            min_answers      = VALUES(min_answers),
            max_answers      = VALUES(max_answers),
            metaphor_density = VALUES(metaphor_density),
            certainty        = VALUES(certainty),
            sentence_length  = VALUES(sentence_length),
            empathy_level    = VALUES(empathy_level),
            notes            = VALUES(notes)
    """)

    try:
        for r in range(2, ws.max_row + 1):
            sid = _cell_int(ws, S_COL_STAGE_ID, r)
            name_kr = _cell_str(ws, S_COL_NAME_KR, r)
            name_en = _cell_str(ws, S_COL_NAME_EN, r)

            if sid is None and name_kr == "" and name_en == "":
                continue

            result.processed += 1

            min_a = _cell_int(ws, S_COL_MIN_ANSWERS, r)
            max_a = _cell_int(ws, S_COL_MAX_ANSWERS, r)
            metaphor = _cell_float(ws, S_COL_METAPHOR, r)
            certainty = _cell_float(ws, S_COL_CERTAINTY, r)
            sent_len = _cell_str(ws, S_COL_SENT_LEN, r)
            empathy = _cell_float(ws, S_COL_EMPATHY, r)
            notes = _cell_str(ws, S_COL_NOTES, r)

            # 필수값 검증
            if sid is None:
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Missing/invalid stage_id (A column)"))
                continue
            if name_kr == "" or name_en == "":
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Missing stage_name_kr(B) or stage_name_en(C)"))
                continue
            if min_a is None or max_a is None:
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Missing min_answers(D) or max_answers(E)"))
                continue
            if metaphor is None or certainty is None or empathy is None:
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Missing metaphor(F) / certainty(G) / empathy(I)"))
                continue
            if sent_len == "":
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Missing sentence_length(H)"))
                continue

            params = {
                "stage_id": sid,
                "stage_name_kr": name_kr,
                "stage_name_en": name_en,
                "min_answers": min_a,
                "max_answers": max_a,
                "metaphor_density": metaphor,
                "certainty": certainty,
                "sentence_length": sent_len,
                "empathy_level": empathy,
                "notes": notes if notes else None,
            }

            res = db.execute(upsert_sql, params)
            if res.rowcount == 1:
                result.inserted += 1
            elif res.rowcount == 2:
                result.updated += 1
            else:
                result.unchanged += 1

        db.commit()
        return result

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {e}")

# =========================
# Idle (혼잣말) Import
# =========================

# 컬럼 매핑 (idle 시트)
# A: id, B: axis_key, C: question_text, D: value, E: enable
I_COL_ID = "A"
I_COL_AXIS_KEY = "B"
I_COL_QUESTION_TEXT = "C"
I_COL_VALUE = "D"
I_COL_ENABLE = "E"

@router.post("/idle/import", response_model=AdminQuestionsImportResponse)
async def admin_idle_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    POST /admin/idle/import
    xlsx 업로드로 psano_idle upsert.
    """

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = _get_sheet(wb, ["idle", "monologue", "혼잣말"]) or wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid xlsx file: {e}")

    result = AdminQuestionsImportResponse()

    upsert_sql = text("""
        INSERT INTO psano_idle
            (id, axis_key, question_text, value, enable)
        VALUES
            (:id, :axis_key, :question_text, :value, :enable)
        ON DUPLICATE KEY UPDATE
            axis_key = VALUES(axis_key),
            question_text = VALUES(question_text),
            value = VALUES(value),
            enable = VALUES(enable)
    """)

    try:
        for r in range(2, ws.max_row + 1):
            id_val = _cell_int(ws, I_COL_ID, r)
            axis_ko = _cell_str(ws, I_COL_AXIS_KEY, r)
            question_text = _cell_str(ws, I_COL_QUESTION_TEXT, r)

            # 빈 행 스킵
            if id_val is None and axis_ko == "" and question_text == "":
                continue

            result.processed += 1

            # axis_key 매핑 (한글 -> 영문)
            axis_key = _map_axis_key(axis_ko)

            # value는 VARCHAR이므로 문자열로 처리
            value = _cell_str(ws, I_COL_VALUE, r) or ""
            enable_raw = _cell_str(ws, I_COL_ENABLE, r)
            enable = _parse_enabled(enable_raw)

            # 필수값 검증
            if id_val is None:
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Invalid or missing ID (A column)"))
                continue

            if not axis_key:
                result.failed += 1
                result.errors.append(ImportErrorItem(
                    row=r, message=f"Invalid axis (B column). Got '{axis_ko}'. Expected: {list(AXIS_MAP.keys())}"
                ))
                continue

            if not question_text:
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Missing question_text (C column)"))
                continue

            if enable is None:
                result.failed += 1
                result.errors.append(ImportErrorItem(
                    row=r, message=f"Invalid enable (E column). Got '{enable_raw}', expected Y/N"
                ))
                continue

            params = {
                "id": id_val,
                "axis_key": axis_key,
                "question_text": question_text,
                "value": value,
                "enable": enable,
            }

            res = db.execute(upsert_sql, params)
            if res.rowcount == 1:
                result.inserted += 1
            elif res.rowcount == 2:
                result.updated += 1
            else:
                result.unchanged += 1

        db.commit()
        return result

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {e}")


# =========================
# Policy Rules Excel Import
# =========================

# 컬럼 매핑 (policy_rules 시트)
# A: 순서(id), B: 주제_한글, C: 키워드, D: 정규식_여부, E: 액션, F: 우선순위, G: 메세지, H: 즉시_완료, I: 활성화
P_COL_ID = "A"
P_COL_CATEGORY = "B"
P_COL_KEYWORDS = "C"
P_COL_IS_REGEX = "D"
P_COL_ACTION = "E"
P_COL_PRIORITY = "F"
P_COL_MESSAGE = "G"
P_COL_SHOULD_END = "H"
P_COL_ENABLED = "I"


@router.post("/policy-rules/import", response_model=AdminQuestionsImportResponse)
async def admin_policy_rules_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    POST /admin/policy-rules/import
    xlsx 업로드로 psano_policy_rules upsert.
    컬럼: 순서, 주제_한글, 키워드, 정규식_여부, 액션, 우선순위, 메세지, 즉시_완료, 활성화
    """

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = _get_sheet(wb, ["policy", "policy_rules", "rules", "정책"]) or wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid xlsx file: {e}")

    result = AdminQuestionsImportResponse()

    # id(순서)로 upsert
    upsert_sql = text("""
        INSERT INTO psano_policy_rules
            (id, category, keywords, is_regex, action, priority, fallback_message, should_end, enabled)
        VALUES
            (:id, :category, :keywords, :is_regex, :action, :priority, :fallback_message, :should_end, :enabled)
        ON DUPLICATE KEY UPDATE
            category = VALUES(category),
            keywords = VALUES(keywords),
            is_regex = VALUES(is_regex),
            action = VALUES(action),
            priority = VALUES(priority),
            fallback_message = VALUES(fallback_message),
            should_end = VALUES(should_end),
            enabled = VALUES(enabled)
    """)

    try:
        for r in range(2, ws.max_row + 1):
            id_val = _cell_int(ws, P_COL_ID, r)
            category = _cell_str(ws, P_COL_CATEGORY, r)
            keywords = _cell_str(ws, P_COL_KEYWORDS, r)
            is_regex_raw = _cell_str(ws, P_COL_IS_REGEX, r)
            action = _cell_str(ws, P_COL_ACTION, r)
            priority_val = _cell_int(ws, P_COL_PRIORITY, r)
            fallback_message = _cell_str(ws, P_COL_MESSAGE, r)
            should_end_raw = _cell_str(ws, P_COL_SHOULD_END, r)
            enabled_raw = _cell_str(ws, P_COL_ENABLED, r)

            # 빈 행 스킵
            if id_val is None and not category and not keywords:
                continue

            result.processed += 1

            # 필수값 검증
            if id_val is None:
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Missing id (A column)"))
                continue

            if not category:
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Missing category (B column)"))
                continue

            if not keywords:
                result.failed += 1
                result.errors.append(ImportErrorItem(row=r, message="Missing keywords (C column)"))
                continue

            # 0/1 파싱
            is_regex = 1 if is_regex_raw in ("1", "Y", "y", "true", "True") else 0
            should_end = 1 if should_end_raw in ("1", "Y", "y", "true", "True") else 0
            enabled = 1 if enabled_raw in ("1", "Y", "y", "true", "True") else 0

            # 기본값 처리
            if not action:
                action = "guide"
            if priority_val is None:
                priority_val = 50

            params = {
                "id": id_val,
                "category": category,
                "keywords": keywords,
                "is_regex": is_regex,
                "action": action,
                "priority": priority_val,
                "fallback_message": fallback_message or "",
                "should_end": should_end,
                "enabled": enabled,
            }

            res = db.execute(upsert_sql, params)
            if res.rowcount == 1:
                result.inserted += 1
            elif res.rowcount == 2:
                result.updated += 1
            else:
                result.unchanged += 1

        db.commit()

        # 캐시 클리어
        try:
            from routers.talk_policy import clear_cache
            clear_cache()
        except Exception:
            pass

        return result

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {e}")


# =========================
# 세션/진행률 조회
# =========================

@router.get("/sessions", response_model=AdminSessionsResponse)
def list_sessions(
    db: Session = Depends(get_db),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    total_row = db.execute(text("SELECT COUNT(*) AS cnt FROM sessions")).mappings().first()
    total = int(total_row["cnt"]) if total_row else 0

    try:
        rows = db.execute(
            text("""
                SELECT id, visitor_name, started_at, ended_at, end_reason
                FROM sessions
                ORDER BY started_at DESC
                LIMIT :limit OFFSET :offset
            """),
            {"limit": limit, "offset": offset}
        ).mappings().all()

        sessions = [{
            "id": int(r["id"]),
            "visitor_name": r["visitor_name"],
            "started_at": iso(r["started_at"]),
            "ended_at": iso(r["ended_at"]),
            "end_reason": r.get("end_reason"),
        } for r in rows]

    except Exception:
        rows = db.execute(
            text("""
                SELECT id, visitor_name, started_at, ended_at
                FROM sessions
                ORDER BY started_at DESC
                LIMIT :limit OFFSET :offset
            """),
            {"limit": limit, "offset": offset}
        ).mappings().all()

        sessions = [{
            "id": int(r["id"]),
            "visitor_name": r["visitor_name"],
            "started_at": iso(r["started_at"]),
            "ended_at": iso(r["ended_at"]),
            "end_reason": None,
        } for r in rows]

    return {"total": total, "sessions": sessions}


@router.get("/progress", response_model=AdminProgressResponse)
def get_progress(db: Session = Depends(get_db)):
    # 설정 로드
    max_questions = get_config(db, "max_questions", MAX_QUESTIONS)
    global_turn_max = get_config(db, "global_turn_max", 365)

    st = db.execute(
        text("SELECT phase, current_question, global_turn_count FROM psano_state WHERE id = 1")
    ).mappings().first()

    if not st:
        raise HTTPException(status_code=500, detail="psano_state(id=1) not found")

    phase = st["phase"]
    if phase not in ("teach", "talk"):
        phase = "teach"

    current_q = int(st["current_question"])
    answered = max_questions if phase == "talk" else max(0, min(max_questions, current_q - 1))
    ratio = float(answered) / float(max_questions) if max_questions > 0 else 0.0
    global_turn_count = int(st.get("global_turn_count") or 0)

    return {
        "phase": phase,
        "current_question": current_q,
        "answered_count": answered,
        "max_questions": max_questions,
        "progress_ratio": ratio,
        "global_turn_count": global_turn_count,
        "global_turn_max": global_turn_max,
    }


# =========================
# 상태 관리 (Reset, Phase, Question)
# =========================

@router.post("/reset", response_model=AdminResetResponse)
def admin_reset(req: AdminResetRequest, db: Session = Depends(get_db)):
    """초기화: state, answers, sessions, personality"""

    def _reset_table(table_name: str):
        db.execute(text(f"DELETE FROM {table_name}"))
        db.execute(text(f"ALTER TABLE {table_name} AUTO_INCREMENT = 1"))

    try:
        ensure_psano_state_row(db)

        if req.reset_answers:
            _reset_table("answers")

        if req.reset_sessions:
            try:
                _reset_table("idle_talk_messages")
            except Exception:
                pass
            _reset_table("sessions")

        if req.reset_personality:
            db.execute(text("""
                UPDATE psano_personality
                SET self_direction = 0, conformity = 0, stimulation = 0, security = 0,
                    hedonism = 0, tradition = 0, achievement = 0, benevolence = 0,
                    power = 0, universalism = 0
                WHERE id = 1
            """))

        if req.reset_state:
            try:
                db.execute(text("""
                    UPDATE psano_state
                    SET phase = 'teach', current_question = 1,
                        formed_at = NULL, persona_prompt = NULL, values_summary = NULL,
                        global_turn_count = 0
                    WHERE id = 1
                """))
            except Exception:
                db.execute(text("""
                    UPDATE psano_state
                    SET phase = 'teach', current_question = 1, global_turn_count = 0
                    WHERE id = 1
                """))

            # 메모리 캐시 동기화
            with LOCK:
                GLOBAL_STATE["phase"] = "teach"
                GLOBAL_STATE["current_question"] = 1
                GLOBAL_STATE["formed_at"] = None
                GLOBAL_STATE["persona_prompt"] = None
                GLOBAL_STATE["values_summary"] = None
                GLOBAL_STATE["global_turn_count"] = 0

        db.commit()
        return AdminResetResponse(
            ok=True,
            reset_answers=req.reset_answers,
            reset_sessions=req.reset_sessions,
            reset_state=req.reset_state,
            reset_personality=req.reset_personality
        )

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"db error: {e}")

@router.post("/phase/set", response_model=AdminPhaseSetResponse)
def admin_set_phase(req: AdminPhaseSetRequest, db: Session = Depends(get_db)):
    """테스트용 phase 강제 변경"""
    if req.phase not in ("teach", "talk"):
        raise HTTPException(status_code=400, detail="invalid phase")

    try:
        ensure_psano_state_row(db)

        if req.phase == "teach":
            try:
                db.execute(text("UPDATE psano_state SET phase = 'teach', formed_at = NULL WHERE id = 1"))
            except Exception:
                db.execute(text("UPDATE psano_state SET phase = 'teach' WHERE id = 1"))

            # 메모리 캐시 동기화
            with LOCK:
                GLOBAL_STATE["phase"] = "teach"
                GLOBAL_STATE["formed_at"] = None
        else:
            formed_at = now_kst_naive()
            try:
                db.execute(
                    text("UPDATE psano_state SET phase = 'talk', formed_at = :formed_at WHERE id = 1"),
                    {"formed_at": formed_at}
                )
            except Exception:
                db.execute(text("UPDATE psano_state SET phase = 'talk' WHERE id = 1"))
                formed_at = None

            # 메모리 캐시 동기화
            with LOCK:
                GLOBAL_STATE["phase"] = "talk"
                GLOBAL_STATE["formed_at"] = formed_at

        db.commit()
        return AdminPhaseSetResponse(ok=True, phase=req.phase)

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"db error: {e}")


@router.post("/state/set_current_question", response_model=AdminSetCurrentQuestionResponse)
def admin_set_current_question(req: AdminSetCurrentQuestionRequest, db: Session = Depends(get_db)):
    """테스트용 현재 질문 강제 설정"""
    try:
        ensure_psano_state_row(db)
        db.execute(
            text("UPDATE psano_state SET current_question = :q WHERE id = 1"),
            {"q": int(req.current_question)}
        )
        db.commit()

        # 메모리 캐시 동기화
        with LOCK:
            GLOBAL_STATE["current_question"] = int(req.current_question)

        return AdminSetCurrentQuestionResponse(ok=True, current_question=int(req.current_question))

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"db error: {e}")

@router.get("/personality", response_model=AdminPersonalityGetResponse)
def admin_personality_get(
    db: Session = Depends(get_db),
):
    """
    GET /admin/personality
    psano_personality(id=1) 10개 축 조회
    """

    try:
        ensure_psano_personality_row(db)

        row = db.execute(
            text("""
                SELECT self_direction, conformity, stimulation, security, hedonism,
                       tradition, achievement, benevolence, power, universalism
                FROM psano_personality
                WHERE id = 1
            """)
        ).mappings().first()

        if not row:
            raise HTTPException(status_code=500, detail="psano_personality(id=1) not found")

        return AdminPersonalityGetResponse(
            self_direction=int(row["self_direction"]),
            conformity=int(row["conformity"]),
            stimulation=int(row["stimulation"]),
            security=int(row["security"]),
            hedonism=int(row["hedonism"]),
            tradition=int(row["tradition"]),
            achievement=int(row["achievement"]),
            benevolence=int(row["benevolence"]),
            power=int(row["power"]),
            universalism=int(row["universalism"]),
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"db error: {e}")


@router.post("/personality/set", response_model=AdminPersonalitySetResponse)
def admin_personality_set(
    req: AdminPersonalitySetRequest,
    db: Session = Depends(get_db),
):
    """
    POST /admin/personality/set
    psano_personality(id=1) 10개 축을 한 번에 덮어쓰기 (테스트용)
    """

    try:
        ensure_psano_personality_row(db)

        res = db.execute(
            text("""
                UPDATE psano_personality
                SET
                    self_direction = :self_direction,
                    conformity     = :conformity,
                    stimulation    = :stimulation,
                    security       = :security,
                    hedonism       = :hedonism,
                    tradition      = :tradition,
                    achievement    = :achievement,
                    benevolence    = :benevolence,
                    power          = :power,
                    universalism   = :universalism
                WHERE id = 1
            """),
            {
                "self_direction": int(req.self_direction),
                "conformity": int(req.conformity),
                "stimulation": int(req.stimulation),
                "security": int(req.security),
                "hedonism": int(req.hedonism),
                "tradition": int(req.tradition),
                "achievement": int(req.achievement),
                "benevolence": int(req.benevolence),
                "power": int(req.power),
                "universalism": int(req.universalism),
            }
        )

        db.commit()
        return AdminPersonalitySetResponse(ok=True, updated=(res.rowcount == 1))

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"db error: {e}")

# =========================
# Persona 생성 (관리자용)
# =========================

@router.post("/generate", response_model=PersonaGenerateResponse)
def admin_persona_generate(req: PersonaGenerateRequest, db: Session = Depends(get_db)):
    """관리자 테스트용: 365 미만이어도 페르소나 생성 가능"""
    try:
        return _generate_persona(db, force=req.force, model=req.model, allow_under_365=True)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"admin persona generate failed: {e}")


# =========================
# Config 관리 (psano_config)
# =========================

@router.get("/config")
def admin_config_list(db: Session = Depends(get_db)):
    """전체 설정 조회"""
    try:
        rows = db.execute(
            text("SELECT config_key, config_value, value_type, description, updated_at FROM psano_config ORDER BY config_key")
        ).mappings().all()

        configs = [{
            "key": r["config_key"],
            "value": r["config_value"],
            "type": r["value_type"],
            "description": r["description"],
            "updated_at": iso(r["updated_at"]),
        } for r in rows]

        return {"configs": configs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"db error: {e}")


@router.put("/config/{key}")
def admin_config_update(
    key: str,
    db: Session = Depends(get_db),
    value: str = Query(...),
    value_type: Optional[str] = Query(None),
):
    """단일 설정 수정"""
    try:
        # 존재 여부 확인
        existing = db.execute(
            text("SELECT config_key FROM psano_config WHERE config_key = :key"),
            {"key": key}
        ).mappings().first()

        if not existing:
            raise HTTPException(status_code=404, detail=f"config not found: {key}")

        if value_type:
            db.execute(
                text("UPDATE psano_config SET config_value = :value, value_type = :vtype WHERE config_key = :key"),
                {"key": key, "value": value, "vtype": value_type}
            )
        else:
            db.execute(
                text("UPDATE psano_config SET config_value = :value WHERE config_key = :key"),
                {"key": key, "value": value}
            )

        db.commit()
        return {"ok": True, "key": key, "value": value}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"db error: {e}")


@router.post("/config/clear-cache")
def admin_config_clear_cache():
    """설정 캐시 초기화"""
    from util.utils import clear_config_cache
    clear_config_cache()
    return {"ok": True, "message": "config cache cleared"}


# =========================
# Prompts 관리 (psano_prompts)
# =========================

@router.get("/prompts")
def admin_prompts_list(db: Session = Depends(get_db)):
    """전체 프롬프트 조회"""
    try:
        rows = db.execute(
            text("SELECT prompt_key, prompt_template, description, updated_at FROM psano_prompts ORDER BY prompt_key")
        ).mappings().all()

        prompts = [{
            "key": r["prompt_key"],
            "template": r["prompt_template"],
            "description": r["description"],
            "updated_at": iso(r["updated_at"]),
        } for r in rows]

        return {"prompts": prompts}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"db error: {e}")


@router.get("/prompts/{key}")
def admin_prompt_get(key: str, db: Session = Depends(get_db)):
    """단일 프롬프트 조회"""
    try:
        row = db.execute(
            text("SELECT prompt_key, prompt_template, description, updated_at FROM psano_prompts WHERE prompt_key = :key"),
            {"key": key}
        ).mappings().first()

        if not row:
            raise HTTPException(status_code=404, detail=f"prompt not found: {key}")

        return {
            "key": row["prompt_key"],
            "template": row["prompt_template"],
            "description": row["description"],
            "updated_at": iso(row["updated_at"]),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"db error: {e}")


@router.put("/prompts/{key}")
def admin_prompt_update(
    key: str,
    db: Session = Depends(get_db),
    template: str = Query(...),
    description: Optional[str] = Query(None),
):
    """단일 프롬프트 수정"""
    try:
        existing = db.execute(
            text("SELECT prompt_key FROM psano_prompts WHERE prompt_key = :key"),
            {"key": key}
        ).mappings().first()

        if not existing:
            raise HTTPException(status_code=404, detail=f"prompt not found: {key}")

        if description is not None:
            db.execute(
                text("UPDATE psano_prompts SET prompt_template = :template, description = :desc WHERE prompt_key = :key"),
                {"key": key, "template": template, "desc": description}
            )
        else:
            db.execute(
                text("UPDATE psano_prompts SET prompt_template = :template WHERE prompt_key = :key"),
                {"key": key, "template": template}
            )

        db.commit()
        return {"ok": True, "key": key}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"db error: {e}")


@router.post("/prompts/clear-cache")
def admin_prompts_clear_cache():
    """프롬프트 캐시 초기화"""
    from util.utils import clear_prompt_cache
    clear_prompt_cache()
    return {"ok": True, "message": "prompt cache cleared"}


# =========================
# Questions 조회/관리
# =========================

@router.get("/questions")
def admin_questions_list(
    db: Session = Depends(get_db),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    enabled_only: bool = Query(False),
):
    """질문 목록 조회"""
    try:
        where_clause = "WHERE enabled = 1" if enabled_only else ""

        total_row = db.execute(
            text(f"SELECT COUNT(*) AS cnt FROM questions {where_clause}")
        ).mappings().first()
        total = int(total_row["cnt"]) if total_row else 0

        rows = db.execute(
            text(f"""
                SELECT id, axis_key, question_text, choice_a, choice_b, enabled, value_a_key, value_b_key
                FROM questions
                {where_clause}
                ORDER BY id ASC
                LIMIT :limit OFFSET :offset
            """),
            {"limit": limit, "offset": offset}
        ).mappings().all()

        questions = [{
            "id": int(r["id"]),
            "axis_key": r["axis_key"],
            "question_text": r["question_text"],
            "choice_a": r["choice_a"],
            "choice_b": r["choice_b"],
            "enabled": bool(r["enabled"]),
            "value_a_key": r["value_a_key"],
            "value_b_key": r["value_b_key"],
        } for r in rows]

        return {"total": total, "questions": questions}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"db error: {e}")


@router.put("/questions/{question_id}/toggle")
def admin_question_toggle(
    question_id: int,
    db: Session = Depends(get_db),
):
    """질문 enabled 토글"""
    try:
        row = db.execute(
            text("SELECT id, enabled FROM questions WHERE id = :id"),
            {"id": question_id}
        ).mappings().first()

        if not row:
            raise HTTPException(status_code=404, detail=f"question not found: {question_id}")

        new_enabled = 0 if row["enabled"] else 1
        db.execute(
            text("UPDATE questions SET enabled = :enabled WHERE id = :id"),
            {"id": question_id, "enabled": new_enabled}
        )
        db.commit()

        return {"ok": True, "id": question_id, "enabled": bool(new_enabled)}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"db error: {e}")


# =========================
# Growth Stages 조회/관리
# =========================

@router.get("/growth-stages")
def admin_growth_stages_list(db: Session = Depends(get_db)):
    """성장단계 목록 조회"""
    try:
        rows = db.execute(
            text("""
                SELECT stage_id, stage_name_kr, stage_name_en, min_answers, max_answers,
                       metaphor_density, certainty, sentence_length, empathy_level, notes, idle_greeting
                FROM psano_growth_stages
                ORDER BY stage_id ASC
            """)
        ).mappings().all()

        stages = [{
            "stage_id": int(r["stage_id"]),
            "stage_name_kr": r["stage_name_kr"],
            "stage_name_en": r["stage_name_en"],
            "min_answers": int(r["min_answers"]),
            "max_answers": int(r["max_answers"]),
            "metaphor_density": float(r["metaphor_density"]) if r["metaphor_density"] else None,
            "certainty": float(r["certainty"]) if r["certainty"] else None,
            "sentence_length": r["sentence_length"],
            "empathy_level": float(r["empathy_level"]) if r["empathy_level"] else None,
            "notes": r["notes"],
            "idle_greeting": r.get("idle_greeting"),
        } for r in rows]

        return {"stages": stages}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"db error: {e}")


@router.put("/growth-stages/{stage_id}")
def admin_growth_stage_update(
    stage_id: int,
    db: Session = Depends(get_db),
    idle_greeting: Optional[str] = Query(None),
    notes: Optional[str] = Query(None),
):
    """성장단계 수정 (idle_greeting, notes)"""
    try:
        row = db.execute(
            text("SELECT stage_id FROM psano_growth_stages WHERE stage_id = :id"),
            {"id": stage_id}
        ).mappings().first()

        if not row:
            raise HTTPException(status_code=404, detail=f"stage not found: {stage_id}")

        updates = []
        params = {"id": stage_id}

        if idle_greeting is not None:
            updates.append("idle_greeting = :idle_greeting")
            params["idle_greeting"] = idle_greeting

        if notes is not None:
            updates.append("notes = :notes")
            params["notes"] = notes

        if not updates:
            return {"ok": True, "message": "nothing to update"}

        db.execute(
            text(f"UPDATE psano_growth_stages SET {', '.join(updates)} WHERE stage_id = :id"),
            params
        )
        db.commit()

        return {"ok": True, "stage_id": stage_id}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"db error: {e}")


# =========================
# Session 상세 (Answers)
# =========================

@router.get("/sessions/{session_id}/answers")
def admin_session_answers(
    session_id: int,
    db: Session = Depends(get_db),
):
    """세션의 답변 목록 조회"""
    try:
        # 세션 정보
        sess = db.execute(
            text("SELECT id, visitor_name, started_at, ended_at, end_reason FROM sessions WHERE id = :id"),
            {"id": session_id}
        ).mappings().first()

        if not sess:
            raise HTTPException(status_code=404, detail=f"session not found: {session_id}")

        # 답변 목록
        rows = db.execute(
            text("""
                SELECT a.id, a.question_id, a.choice, a.chosen_value_key, a.assistant_reaction, a.created_at,
                       q.question_text, q.choice_a, q.choice_b
                FROM answers a
                LEFT JOIN questions q ON a.question_id = q.id
                WHERE a.session_id = :sid
                ORDER BY a.id ASC
            """),
            {"sid": session_id}
        ).mappings().all()

        answers = [{
            "id": int(r["id"]),
            "question_id": int(r["question_id"]),
            "question_text": r["question_text"],
            "choice": r["choice"],
            "choice_text": r["choice_a"] if r["choice"] == "A" else r["choice_b"],
            "chosen_value_key": r["chosen_value_key"],
            "assistant_reaction": r["assistant_reaction"],
            "created_at": iso(r["created_at"]),
        } for r in rows]

        return {
            "session": {
                "id": int(sess["id"]),
                "visitor_name": sess["visitor_name"],
                "started_at": iso(sess["started_at"]),
                "ended_at": iso(sess["ended_at"]),
                "end_reason": sess.get("end_reason"),
            },
            "answers": answers,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"db error: {e}")


# =========================
# Talk Topics 조회
# =========================

@router.get("/topics")
def admin_topics_list(db: Session = Depends(get_db)):
    """대화 주제 목록 조회"""
    try:
        rows = db.execute(
            text("SELECT id, title, description FROM talk_topics ORDER BY id ASC")
        ).mappings().all()

        topics = [{
            "id": int(r["id"]),
            "title": r["title"],
            "description": r["description"],
        } for r in rows]

        return {"topics": topics}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"db error: {e}")


# =========================
# 현재 Persona 조회
# =========================

@router.get("/persona")
def admin_persona_get(db: Session = Depends(get_db)):
    """현재 생성된 persona_prompt 및 values_summary 조회"""
    try:
        row = db.execute(
            text("""
                SELECT phase, persona_prompt, values_summary, formed_at, current_question
                FROM psano_state
                WHERE id = 1
            """)
        ).mappings().first()

        if not row:
            raise HTTPException(status_code=500, detail="psano_state(id=1) not found")

        return {
            "phase": row["phase"],
            "persona_prompt": row["persona_prompt"],
            "values_summary": row["values_summary"],
            "formed_at": iso(row["formed_at"]),
            "current_question": int(row["current_question"]) if row["current_question"] else None,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"db error: {e}")


# =========================
# Quick Test (자동 답변 제출)
# =========================

@router.post("/quick-test")
def admin_quick_test(
    db: Session = Depends(get_db),
    visitor_name: str = Query("QuickTest", description="테스트 방문자 이름"),
    answer_count: int = Query(5, ge=1, le=10, description="자동 제출할 답변 수"),
):
    """
    Quick Test: 세션 시작 → 지정된 수의 답변 자동 제출 → 세션 종료
    각 질문에 랜덤(A/B)으로 응답
    """
    import random

    try:
        # 1) psano_state에서 현재 질문 조회
        st = db.execute(
            text("SELECT current_question FROM psano_state WHERE id = 1")
        ).mappings().first()

        if not st:
            raise HTTPException(status_code=500, detail="psano_state(id=1) not found")

        start_question_id = int(st["current_question"])

        # 2) 세션 생성
        from util.utils import now_kst_naive
        started_at = now_kst_naive()

        result = db.execute(
            text("""
                INSERT INTO sessions (visitor_name, started_at, start_question_id)
                VALUES (:name, :started_at, :start_qid)
            """),
            {"name": visitor_name, "started_at": started_at, "start_qid": start_question_id}
        )
        session_id = int(getattr(result, "lastrowid", 0) or 0)

        if not session_id:
            raise HTTPException(status_code=500, detail="failed to create session")

        # 3) 답변 자동 제출
        answers_submitted = []
        current_qid = start_question_id

        for i in range(answer_count):
            # 질문 조회
            q = db.execute(
                text("""
                    SELECT id, value_a_key, value_b_key, enabled
                    FROM questions
                    WHERE id >= :qid AND enabled = 1
                    ORDER BY id ASC
                    LIMIT 1
                """),
                {"qid": current_qid}
            ).mappings().first()

            if not q:
                break  # 더 이상 질문이 없음

            qid = int(q["id"])
            choice = random.choice(["A", "B"])
            chosen_value_key = q["value_a_key"] if choice == "A" else q["value_b_key"]

            # 답변 저장
            db.execute(
                text("""
                    INSERT INTO answers (session_id, question_id, choice, chosen_value_key)
                    VALUES (:sid, :qid, :choice, :value_key)
                """),
                {"sid": session_id, "qid": qid, "choice": choice, "value_key": chosen_value_key}
            )

            answers_submitted.append({
                "question_id": qid,
                "choice": choice,
                "chosen_value_key": chosen_value_key,
            })

            current_qid = qid + 1

        # 4) psano_personality 업데이트
        for ans in answers_submitted:
            col = ans["chosen_value_key"]
            if col:
                db.execute(
                    text(f"UPDATE psano_personality SET `{col}` = `{col}` + 1 WHERE id = 1"),
                )

        # 5) psano_state.current_question 업데이트
        if answers_submitted:
            next_q = db.execute(
                text("""
                    SELECT id FROM questions
                    WHERE id > :last_qid AND enabled = 1
                    ORDER BY id ASC
                    LIMIT 1
                """),
                {"last_qid": answers_submitted[-1]["question_id"]}
            ).mappings().first()

            new_current_q = int(next_q["id"]) if next_q else 381
            db.execute(
                text("UPDATE psano_state SET current_question = :q WHERE id = 1"),
                {"q": new_current_q}
            )

        # 6) 세션 종료
        ended_at = now_kst_naive()
        db.execute(
            text("""
                UPDATE sessions
                SET ended_at = :ended_at, end_reason = 'completed'
                WHERE id = :sid
            """),
            {"ended_at": ended_at, "sid": session_id}
        )

        db.commit()

        return {
            "ok": True,
            "session_id": session_id,
            "visitor_name": visitor_name,
            "answers_count": len(answers_submitted),
            "answers": answers_submitted,
            "start_question_id": start_question_id,
            "next_question_id": new_current_q if answers_submitted else start_question_id,
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"quick test failed: {e}")


@router.get("/idle/list")
def admin_idle_list(
    db: Session = Depends(get_db),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    enabled_only: bool = Query(False),
):
    """
    GET /admin/idle/list
    psano_idle 목록 조회
    """

    try:
        where_clause = "WHERE enable = 1" if enabled_only else ""

        total_row = db.execute(
            text(f"SELECT COUNT(*) AS cnt FROM psano_idle {where_clause}")
        ).mappings().first()
        total = int(total_row["cnt"]) if total_row else 0

        rows = db.execute(
            text(f"""
                SELECT id, axis_key, question_text, value, enable
                FROM psano_idle
                {where_clause}
                ORDER BY id ASC
                LIMIT :limit OFFSET :offset
            """),
            {"limit": limit, "offset": offset}
        ).mappings().all()

        items = [{
            "id": int(r["id"]),
            "axis_key": r["axis_key"],
            "question_text": r["question_text"],
            "value": r["value"] or "",
            "enable": bool(r["enable"]),
        } for r in rows]

        return {"total": total, "items": items}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"db error: {e}")


@router.put("/idle/{idle_id}/toggle")
def admin_idle_toggle(
    idle_id: int,
    db: Session = Depends(get_db),
):
    """
    PUT /admin/idle/{idle_id}/toggle
    enable 토글
    """

    try:
        row = db.execute(
            text("SELECT id, enable FROM psano_idle WHERE id = :id"),
            {"id": idle_id}
        ).mappings().first()

        if not row:
            raise HTTPException(status_code=404, detail=f"idle not found: {idle_id}")

        new_enable = 0 if row["enable"] else 1
        db.execute(
            text("UPDATE psano_idle SET enable = :enable WHERE id = :id"),
            {"id": idle_id, "enable": new_enable}
        )
        db.commit()

        return {"ok": True, "id": idle_id, "enable": bool(new_enable)}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"db error: {e}")


# ──────────────────────────────────────────────────────────────
# Policy Rules 관리
# ──────────────────────────────────────────────────────────────

@router.get("/policy-rules")
def admin_policy_rules_list(
    db: Session = Depends(get_db),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    enabled_only: bool = Query(False),
):
    """
    GET /admin/policy-rules
    psano_policy_rules 목록 조회
    """
    try:
        where_clause = "WHERE enabled = TRUE" if enabled_only else ""

        total_row = db.execute(
            text(f"SELECT COUNT(*) AS cnt FROM psano_policy_rules {where_clause}")
        ).mappings().first()
        total = int(total_row["cnt"]) if total_row else 0

        rows = db.execute(
            text(f"""
                SELECT id, category, keywords, is_regex, action, priority,
                       fallback_message, should_end, enabled
                FROM psano_policy_rules
                {where_clause}
                ORDER BY priority DESC, id ASC
                LIMIT :limit OFFSET :offset
            """),
            {"limit": limit, "offset": offset}
        ).mappings().all()

        items = [{
            "id": int(r["id"]),
            "category": r["category"] or "",
            "keywords": r["keywords"] or "",
            "is_regex": bool(r["is_regex"]),
            "action": r["action"] or "redirect",
            "priority": int(r["priority"] or 50),
            "fallback_message": r["fallback_message"] or "",
            "should_end": bool(r["should_end"]),
            "enabled": bool(r["enabled"]),
        } for r in rows]

        return {"total": total, "items": items}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"db error: {e}")


@router.put("/policy-rules/{rule_id}/toggle")
def admin_policy_rule_toggle(
    rule_id: int,
    db: Session = Depends(get_db),
):
    """
    PUT /admin/policy-rules/{rule_id}/toggle
    enabled 토글
    """
    try:
        row = db.execute(
            text("SELECT id, enabled FROM psano_policy_rules WHERE id = :id"),
            {"id": rule_id}
        ).mappings().first()

        if not row:
            raise HTTPException(status_code=404, detail=f"policy rule not found: {rule_id}")

        new_enabled = not bool(row["enabled"])
        db.execute(
            text("UPDATE psano_policy_rules SET enabled = :enabled WHERE id = :id"),
            {"id": rule_id, "enabled": new_enabled}
        )
        db.commit()

        # 캐시 클리어 (talk_policy에서 import)
        from routers.talk_policy import clear_cache
        clear_cache()

        return {"ok": True, "id": rule_id, "enabled": new_enabled}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"db error: {e}")